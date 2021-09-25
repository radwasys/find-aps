from selenium import webdriver
from openpyxl import Workbook, load_workbook

load = input('do you already have a apartements.xlsx file ?[yes/no]')

if load.lower() == 'no':
    wb = Workbook()
    
if load.lower() == 'yes':
    wb = load_workbook('apartements.xlsx')

else:
    raise Exception('response not valid')

try:
    ws = wb['Bayut website']
except:
    ws = wb.create_sheet('Bayut website')


PATH = 'C:/Program Files (x86)/chromedriver'
option = webdriver.ChromeOptions()
option.add_argument('headless')
driver = webdriver.Chrome(PATH, options=option)

driver.get('https://www.bayut.sa/en/western-region/apartments-for-rent-in-makkah/?residence_type=family')

prices = driver.find_elements_by_class_name('f343d9ce')
frequencies = driver.find_elements_by_class_name('e76c7aca')
locations = driver.find_elements_by_class_name('_7afabd84')
bedrooms = driver.find_elements_by_css_selector('span.b6a29bc0[aria-label="Beds"]')
bathrooms = driver.find_elements_by_css_selector('span.b6a29bc0[aria-label="Baths"]')
areas = driver.find_elements_by_css_selector('span.b6a29bc0[aria-label="Area"]')
links = driver.find_elements_by_class_name('_287661cb') #link.get_attribute('href')

x = 0

for index, price in enumerate(prices):
    location = locations[index].text
    number_of_bedrooms = int(bedrooms[index].text)
    number_of_bathrooms = int(bathrooms[index].text)
    if 'Al Nasim' in location and number_of_bedrooms > 2 and number_of_bathrooms >= 2:
        x += 1
        price = price.text.split(',')
        price = int(price[0] + price[1])
        frequency = frequencies[index].text
        area = int(areas[index].text.split(' ')[0])
        link = links[index].get_attribute('href')

        ws.append(['apartement number', 'Price', 'Location', 'Number of bedrooms', 'Number of bathrooms', 'Area', 'Links'])
        ws.append([x, price, location, number_of_bedrooms, number_of_bathrooms, area, link])
    


wb.save("apartements.xlsx")

driver.close()