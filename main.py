from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook
import time
import os

if os.path.isfile('apartements.xlsx'):
    wb = load_workbook('apartements.xlsx')
    clear = input('already have apartements file, clear it? [yes,no]')
    if clear.lower() == 'yes':
        ws = wb.active
        for i in range(100):
            ws.append(['','','','','','','','','',''])
else:
    wb = Workbook()

ws = wb.active
ws.append(['apartement number', 'price', 'location', 'number of bedrooms', 'number of bathrooms', 'area'])

PATH = 'C:/Program Files (x86)/chromedriver'
# option = webdriver.ChromeOptions()
# option.add_argument('headless')
driver = webdriver.Chrome(PATH)

driver.get('https://www.propertyfinder.sa/en/search?bf=3&btf=2&c=2&fu=3&l=11&ob=pa&page=1&rp=y&t=1')

# number_of_pages = driver.find_element_by_class_name('pagination__links').text.split('\n')[-1]


list_of_flats = driver.find_elements_by_class_name('card-list__item')[:25]
list_of_prices = driver.find_elements_by_class_name('card__price')
locations = driver.find_elements_by_class_name('card__location-text')
bedrooms = driver.find_elements_by_class_name('card__property-amenity--bedrooms')
bathrooms = driver.find_elements_by_class_name('card__property-amenity--bathrooms')
areas = driver.find_elements_by_class_name('card__property-amenity--area')

number_of_flats = len(list_of_flats)
prices = []
for price in list_of_prices:
    price_list = price.text.split(' ')[0].split(',')
    prices.append(price_list[0]+price_list[1])

print(len(prices), len(bedrooms), len(bathrooms))

for index, price in enumerate(prices):
    location = locations[index].text
    if int(price) <= 40000:
        number_of_bedrooms = bedrooms[index].text
        number_of_bathrooms = bathrooms[index].text
        area = areas[index].text

        ws.append([index+1, price, location, number_of_bedrooms, number_of_bathrooms, area])


        # print(f'---------aprtement number {index+1}----------')
        # print(f'PRICE: {price}')
        # print(f'LOCATION: {location}')
        # print(f'NUMBER OF BEDROOMS: {number_of_bedrooms}')
        # print(f'NUMBER OF BATHROOMS: {number_of_bathrooms}')
        # print(f'AREA OF THE APARTEMENT: {area}')
        # print('--------------------------------------------')


    
wb.save('apartements.xlsx')
driver.quit()