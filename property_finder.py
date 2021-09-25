from selenium import webdriver
from openpyxl import Workbook, load_workbook

load = input('do you already have a apartements.xlsx file ?[yes/no]')
if load.lower() == 'yes':
    wb = load_workbook('apartements.xlsx')
    
if load.lower() == 'no':
    wb = Workbook()
else:
    raise Exception("response not valid")

try:
    ws = wb['Property finder website']
except:
    ws = wb.create_sheet('Property finder website')
ws.append(['apartement number', 'price', 'location', 'number of bedrooms', 'number of bathrooms', 'area', 'frequencies', 'links'])

PATH = 'C:/Program Files (x86)/chromedriver'
option = webdriver.ChromeOptions()
option.add_argument('headless')
driver = webdriver.Chrome(PATH, options=option)

driver.get('https://www.propertyfinder.sa/en/search?bf=3&btf=2&c=2&fu=3&l=11&ob=pa&page=1&rp=y&t=1')

prices = driver.find_elements_by_class_name('card__price')
locations = driver.find_elements_by_class_name('card__location-text')
bedrooms = driver.find_elements_by_class_name('card__property-amenity--bedrooms')
bathrooms = driver.find_elements_by_class_name('card__property-amenity--bathrooms')
areas = driver.find_elements_by_class_name('card__property-amenity--area')
links = driver.find_elements_by_class_name('card--clickable')

x = 0
for index, price in enumerate(prices):
    price = price.text.split(' ')[0].split(',')
    price = int(price[0] + price[1])
    location = locations[index].text
    number_of_bedrooms = int(bedrooms[index].text)
    number_of_bathrooms = int(bathrooms[index].text)
    if price <= 40000 and number_of_bathrooms >= 2 and number_of_bedrooms > 2 and 'Jeddah' not in location:
        x += 1
        area = areas[index].text
        link = links[index].get_attribute('href')
        ws.append([x, price, location, number_of_bedrooms, number_of_bathrooms, area, 'Yearly', link])



    
wb.save('apartements.xlsx')
driver.quit()